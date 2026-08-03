# core/excel_utils.py
"""
Excelファイル読み込みまわりの補助的なチェック処理をまとめたモジュール。

pandas.read_excel (openpyxlエンジン) は、数式セルについては原則として
Excel自身が最後に保存した「計算済みキャッシュ値」を読み取る。
そのため、Excel以外のツールで数式だけを書き込んで保存したファイルや、
手動計算モードのまま保存されたファイルでは、キャッシュ値が存在せず
セルが空 (None/NaN) として読み込まれてしまうことがある。
これは見た目には「データが欠損している」ようにしか見えず気づきにくいため、
読み込み前に検出して警告できるようにする。
"""
import logging

import openpyxl

logger = logging.getLogger(__name__)


def find_unevaluated_formula_cells(file_path, sheet_name=None, max_examples=5, max_scan_cells=200_000):
    """
    指定したExcelファイル(のシート)内で、数式セルであるにもかかわらず
    計算済みの値を持たない(data_only=Trueで読んでも None になる)セルを探す。

    大きなファイルでの負荷を避けるため、走査するセル数に上限を設けており、
    上限に達した時点でその時点までの結果を返す(scanned_all=False)。

    Args:
        file_path (str): Excelファイルのパス。
        sheet_name (str, optional): 対象シート名。Noneなら全シートを対象にする。
        max_examples (int): 収集する具体例(シート名!セル番地)の最大数。
        max_scan_cells (int): 走査するセルの最大数。

    Returns:
        tuple (bool, list[str], bool): (見つかったか, 具体例のリスト, 全体を走査しきったか)
    """
    wb_formulas = None
    wb_values = None
    try:
        # ★ ファイルが存在しない/壊れている場合、load_workbook自体が例外を送出しうる。
        # この関数は「検査に失敗したら安全側 (見つからなかった扱い) に倒れる」契約
        # なので、読み込み自体もtry節の中に含める。
        wb_formulas = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
        wb_values = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

        sheets = [sheet_name] if sheet_name else wb_formulas.sheetnames
        examples = []
        scanned = 0

        for sname in sheets:
            if sname not in wb_formulas.sheetnames or sname not in wb_values.sheetnames:
                continue
            ws_formulas = wb_formulas[sname]
            ws_values = wb_values[sname]

            for row_f, row_v in zip(ws_formulas.iter_rows(), ws_values.iter_rows()):
                for cell_f, cell_v in zip(row_f, row_v):
                    scanned += 1
                    if cell_f.data_type == 'f' and cell_v.value is None:
                        examples.append(f"{sname}!{cell_f.coordinate}")
                        if len(examples) >= max_examples:
                            return True, examples, (scanned < max_scan_cells)
                    if scanned >= max_scan_cells:
                        return bool(examples), examples, False

        return bool(examples), examples, True
    except Exception:
        logger.exception("数式セルの検査中にエラーが発生しました: %s", file_path)
        return False, [], True
    finally:
        if wb_formulas is not None:
            wb_formulas.close()
        if wb_values is not None:
            wb_values.close()
