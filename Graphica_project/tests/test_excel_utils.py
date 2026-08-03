# tests/test_excel_utils.py
"""core/excel_utils.py (数式セルの値ずれ検出) に対するテスト。"""
import openpyxl
import pandas as pd
import pytest

from core.excel_utils import find_unevaluated_formula_cells


def test_detects_formula_cell_without_cached_value(tmp_path):
    path = tmp_path / "formula.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws['A1'] = 'x'
    ws['B1'] = 'y'
    ws['A2'] = 1
    ws['B2'] = '=A2*2'  # openpyxlで書き込んだ数式はキャッシュ値を持たない
    ws['A3'] = 2
    ws['B3'] = 20
    wb.save(path)

    found, examples, scanned_all = find_unevaluated_formula_cells(str(path), 'Sheet1')

    assert found is True
    assert any('B2' in ex for ex in examples)
    assert scanned_all is True


def test_no_false_positive_on_clean_file(tmp_path):
    path = tmp_path / "clean.xlsx"
    pd.DataFrame({'x': [1, 2, 3], 'y': [10, 20, 30]}).to_excel(path, index=False, sheet_name='Sheet1')

    found, examples, _ = find_unevaluated_formula_cells(str(path), 'Sheet1')

    assert found is False
    assert examples == []


def test_examples_capped_at_max_examples(tmp_path):
    path = tmp_path / "many_formulas.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in range(1, 11):
        ws[f'A{row}'] = f'=B{row}*2'  # 10個すべて未計算の数式セル

    wb.save(path)

    found, examples, scanned_all = find_unevaluated_formula_cells(str(path), 'Sheet1', max_examples=3)

    assert found is True
    assert len(examples) == 3


def test_sheet_name_scoping(tmp_path):
    """sheet_name を指定した場合、そのシートだけを検査する"""
    path = tmp_path / "two_sheets.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Clean"
    ws1['A1'] = 1
    ws2 = wb.create_sheet("HasFormula")
    ws2['A1'] = '=1+1'
    wb.save(path)

    found_clean, _, _ = find_unevaluated_formula_cells(str(path), 'Clean')
    found_formula, _, _ = find_unevaluated_formula_cells(str(path), 'HasFormula')

    assert found_clean is False
    assert found_formula is True


def test_nonexistent_file_does_not_raise(tmp_path):
    """
    ファイルが存在しない/壊れている場合でも例外を伝播させず、
    "見つからなかった" 扱いで安全側に倒れることを確認する
    (呼び出し側の main_window.py は、この関数の失敗で読み込みフロー全体を
    止めないことを前提にしている)。
    """
    missing_path = tmp_path / "does_not_exist.xlsx"
    found, examples, scanned_all = find_unevaluated_formula_cells(str(missing_path), 'Sheet1')
    assert found is False
    assert examples == []
